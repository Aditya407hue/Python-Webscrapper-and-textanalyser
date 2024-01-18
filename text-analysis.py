import openpyxl
import nltk
import regex as re
from nltk.corpus import stopwords

# Download NLTK resources
nltk.download('punkt')
nltk.download('stopwords')

# Function to perform sentimental analysis and calculate derived variables
def perform_sentimental_analysis(text):

    stop_words = set(nltk.corpus.stopwords.words('english'))

    # Load Master Dictionary of Positive and Negative words
    positive_words = set()
    negative_words = set()

    with open('MasterDictionary/positive-words.txt', 'r') as file:
        positive_words = set(file.read().splitlines())
    with open('MasterDictionary/negative-words.txt', 'r') as file:
        negative_words = set(file.read().splitlines())

    # Clean text using Stop Words Lists
    cleaned_words = [word.lower() for word in nltk.word_tokenize(text) if
                     word.isalnum() and word.lower() not in stop_words]

    # Calculate derived variables
    positive_score = sum(1 for word in cleaned_words if word in positive_words)
    negative_score = sum(1 for word in cleaned_words if word in negative_words)

    # Ensure that there are no division by zero errors
    total_words_after_cleaning = len(cleaned_words) + 0.000001
    positive_score += 0.000001
    negative_score += 0.000001

    polarity_score = (positive_score - negative_score) / (positive_score + negative_score)
    subjectivity_score = (positive_score + negative_score) / total_words_after_cleaning

    return positive_score, negative_score, polarity_score, subjectivity_score


# Function to perform analysis of readability
def perform_readability_analysis(text):
    # Calculate Average Sentence Length
    sentences = nltk.sent_tokenize(text)
    words = nltk.word_tokenize(text)
    avg_sentence_length = len(words) / len(sentences) if len(sentences) > 0 else 0

    # Calculate Percentage of Complex Words
    complex_words = [word for word in words if
                     len(word) > 2]  # Adjust the definition of complex words based on your requirements
    percentage_complex_words: float | int = len(complex_words) / len(words) if len(words) > 0 else 0

    # Calculate Fog Index
    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

    return avg_sentence_length, percentage_complex_words, fog_index


# Function to perform analysis of other variables
def average_number_of_words_per_sentence(text):
    sentences = nltk.sent_tokenize(text)
    words = nltk.word_tokenize(text)
    return len(words) / len(sentences) if len(sentences) > 0 else 0


def complex_word_count(text):
    words = nltk.word_tokenize(text)
    return sum(
        1 for word in words if len(word) > 2)  # Adjust the definition of complex words based on your requirements


def word_count(text):
    stop_words = set(stopwords.words('english'))
    words = [word.lower() for word in nltk.word_tokenize(text) if word.isalnum() and word.lower() not in stop_words]
    return len(words)


def syllable_count_per_word(text):
    def count_syllables(word):
        # Simple syllable counting algorithm
        vowels = "aeiouy"
        count = 0
        is_prev_vowel = False
        for char in word.lower():
            if char in vowels:
                if not is_prev_vowel:
                    count += 1
                is_prev_vowel = True
            else:
                is_prev_vowel = False
        return count

    words = nltk.word_tokenize(text)
    return sum(count_syllables(word) for word in words)


def personal_pronouns_count(text):
    pronoun_list = ["I", "we", "my", "ours", "us"]
    # Regex to match personal pronouns (case-insensitive)
    pronoun_pattern = re.compile(r'\b(?:' + '|'.join(re.escape(pronoun) for pronoun in pronoun_list) + r')\b',
                                 re.IGNORECASE)
    return len(re.findall(pronoun_pattern, text))


def average_word_length(text):
    words = nltk.word_tokenize(text)
    total_chars = sum(len(word) for word in words)
    return total_chars / len(words) if len(words) > 0 else 0

# Load URLs from the Excel file
workbook_input = openpyxl.load_workbook('input.xlsx')
sheet_input = workbook_input.active

# Load the output structure Excel file for writing results
workbook_output = openpyxl.load_workbook('Output Data Structure.xlsx')
sheet_output = workbook_output.active

# Iterate through each row in the Excel input file
for row_input, row_output in zip(sheet_input.iter_rows(min_row=2, values_only=True), sheet_output.iter_rows(min_row=2)):
    url_id, filename = row_input

    try:
        # Read article text from the corresponding text file
        with open(f'{url_id}.txt', 'r', encoding='utf-8') as file:
            article_text: str = file.read()
        # Perform sentimental analysis
        positive_score, negative_score, polarity_score, subjectivity_score = perform_sentimental_analysis(article_text)

        # Update the output structure Excel file with computed variables
        row_output[2].value = positive_score
        row_output[3].value = negative_score
        row_output[4].value = polarity_score
        row_output[5].value = subjectivity_score

        # Perform analysis of readability
        avg_sentence_length, percentage_complex_words, fog_index = perform_readability_analysis(article_text)

        # Update the output structure Excel file with computed variables
        row_output[6].value = avg_sentence_length
        row_output[7].value = percentage_complex_words
        row_output[8].value = fog_index

        avg_words_per_sentence = average_number_of_words_per_sentence(article_text)
        complex_words_count = complex_word_count(article_text)
        total_words_count = word_count(article_text)
        syllable_count_per_word_value = syllable_count_per_word(article_text)
        personal_pronouns_count_value = personal_pronouns_count(article_text)
        avg_word_length_value = average_word_length(article_text)

        row_output[9].value = avg_words_per_sentence
        row_output[10].value = complex_words_count
        row_output[11].value = total_words_count
        row_output[12].value = syllable_count_per_word_value
        row_output[13].value = personal_pronouns_count_value
        row_output[14].value = avg_word_length_value

        print(f'Successfully analyzed text for {url_id}')

    except Exception as e:
        print(f'Error processing {url_id}: {e}')

# Save the results in the output structure Excel file
workbook_output.save('Output Data Structure.xlsx')

# Close the Excel files
workbook_input.close()
workbook_output.close()
